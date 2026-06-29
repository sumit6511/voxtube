"""
Export module for VoxTube.

Generates a multi-sheet Excel workbook (.xlsx) summarizing a completed
analysis job — suitable for sharing with stakeholders who don't have
access to the live dashboard, or for offline record-keeping.

Sheets:
  1. Summary  - job metadata, sentiment / language / toxicity breakdowns
  2. Comments - one row per comment with all NLP annotations
  3. Topics   - discovered topics with keywords and sentiment distribution
"""

from __future__ import annotations

import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared styles ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1E2330", end_color="1E2330", fill_type="solid")
HEADER_FONT = Font(bold=True, color="F59E0B", size=11)
TITLE_FONT  = Font(bold=True, size=14)
LABEL_FONT  = Font(bold=True, size=10, color="6B7280")
THIN_BORDER = Border(bottom=Side(style="thin", color="2D3446"))

CENTER = Alignment(horizontal="center", vertical="center")
WRAP   = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _autosize_columns(ws, widths: dict[int, int]):
    """widths: {column_index (1-based): width}"""
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _safe_scores(json_str: str | None) -> dict[str, float]:
    try:
        return json.loads(json_str) if json_str else {}
    except Exception:
        return {}


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, job, comments: list, topics: list):
    ws = wb.active
    ws.title = "Summary"

    total = len(comments)

    # Sentiment counts
    sent_counts = {"positive": 0, "neutral": 0, "negative": 0}
    for c in comments:
        if c.sentiment_label in sent_counts:
            sent_counts[c.sentiment_label] += 1

    # Language counts
    lang_counts = {"nepali": 0, "english": 0, "neplish": 0}
    for c in comments:
        lang = c.lang if c.lang in lang_counts else "neplish"
        lang_counts[lang] += 1

    # Toxicity
    toxic_count = sum(1 for c in comments if c.is_toxic)

    row = 1

    # Title
    ws.cell(row=row, column=1, value="VoxTube Analysis Report").font = TITLE_FONT
    row += 2

    # Job metadata
    meta_rows = [
        ("Video Title", job.video_title or "—"),
        ("Video URL",   job.youtube_url),
        ("Job ID",      job.id),
        ("Total Comments", total),
        ("Topics Discovered", len(topics)),
    ]
    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    # Sentiment breakdown
    ws.cell(row=row, column=1, value="Sentiment Breakdown").font = Font(bold=True, size=12)
    row += 1
    _style_header_row(ws, row, 3)
    ws.cell(row=row, column=1, value="Sentiment")
    ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=3, value="Percentage")
    row += 1
    for label in ["positive", "neutral", "negative"]:
        count = sent_counts[label]
        pct   = (count / total * 100) if total else 0
        ws.cell(row=row, column=1, value=label.capitalize())
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        row += 1
    row += 1

    # Language breakdown
    ws.cell(row=row, column=1, value="Language Breakdown").font = Font(bold=True, size=12)
    row += 1
    _style_header_row(ws, row, 3)
    ws.cell(row=row, column=1, value="Language")
    ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=3, value="Percentage")
    row += 1
    lang_labels = {"nepali": "Nepali", "english": "English", "neplish": "Neplish"}
    for key, label in lang_labels.items():
        count = lang_counts[key]
        pct   = (count / total * 100) if total else 0
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        row += 1
    row += 1

    # Toxicity summary
    ws.cell(row=row, column=1, value="Toxicity Summary").font = Font(bold=True, size=12)
    row += 1
    _style_header_row(ws, row, 3)
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=3, value="Percentage")
    row += 1
    tox_pct = (toxic_count / total * 100) if total else 0
    ws.cell(row=row, column=1, value="Flagged as toxic")
    ws.cell(row=row, column=2, value=toxic_count)
    ws.cell(row=row, column=3, value=f"{tox_pct:.1f}%")
    row += 1

    _autosize_columns(ws, {1: 26, 2: 50, 3: 14})


def _build_comments_sheet(wb: Workbook, comments: list):
    ws = wb.create_sheet("Comments")

    headers = [
        "#", "Comment Text", "Language", "Sentiment (XLM-RoBERTa)",
        "Sentiment Score", "VADER Label", "VADER Compound",
        "Topic ID", "Is Toxic", "Toxicity Categories",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, c in enumerate(comments, start=1):
        scores = _safe_scores(c.toxicity_json)
        flagged = [k for k, v in scores.items() if v >= 0.5]

        ws.append([
            i,
            c.original_text,
            c.lang or "—",
            c.sentiment_label or "—",
            round(c.sentiment_score, 4) if c.sentiment_score is not None else "—",
            c.vader_label or "—",
            round(c.vader_compound, 4) if c.vader_compound is not None else "—",
            c.topic_id if c.topic_id is not None and c.topic_id != -1 else "—",
            "Yes" if c.is_toxic else "No",
            ", ".join(flagged) if flagged else "—",
        ])

        # Wrap the comment text cell
        ws.cell(row=i + 1, column=2).alignment = WRAP

    _autosize_columns(ws, {
        1: 5, 2: 60, 3: 10, 4: 20, 5: 14, 6: 12, 7: 14, 8: 9, 9: 9, 10: 30,
    })

    ws.freeze_panes = "A2"   # keep header row visible while scrolling


def _build_topics_sheet(wb: Workbook, topics: list):
    ws = wb.create_sheet("Topics")

    headers = [
        "Topic ID", "Label", "Keywords", "Comment Count",
        "Positive", "Neutral", "Negative", "Dominant Sentiment",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for t in topics:
        try:
            keywords = json.loads(t.keywords_json) if t.keywords_json else []
        except Exception:
            keywords = []

        counts = {
            "Positive": t.positive_count,
            "Neutral":  t.neutral_count,
            "Negative": t.negative_count,
        }
        dominant = max(counts, key=counts.get) if t.comment_count else "—"

        ws.append([
            t.topic_id,
            t.label or "—",
            ", ".join(keywords),
            t.comment_count,
            t.positive_count,
            t.neutral_count,
            t.negative_count,
            dominant,
        ])
        ws.cell(row=ws.max_row, column=3).alignment = WRAP

    _autosize_columns(ws, {
        1: 9, 2: 24, 3: 50, 4: 14, 5: 10, 6: 10, 7: 10, 8: 18,
    })

    ws.freeze_panes = "A2"


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_excel_report(job, comments: list, topics: list) -> io.BytesIO:
    """
    Build a 3-sheet Excel workbook for a completed analysis job.

    Args:
        job:      Job ORM instance (must have status == 'done')
        comments: list of Comment ORM instances for this job
        topics:   list of Topic ORM instances for this job

    Returns:
        io.BytesIO containing the .xlsx file, seeked to position 0,
        ready to be streamed in a FastAPI response.
    """
    wb = Workbook()

    _build_summary_sheet(wb, job, comments, topics)
    _build_comments_sheet(wb, comments)
    _build_topics_sheet(wb, topics)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
